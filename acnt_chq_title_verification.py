import openpyxl
import re


class title_verification:
    acnt_title = ""
    cheque_title = ""
    ismatched = False
    dict_flag = False
    status = ""
    reason = " "
    modified_status, success_status, Unsuccess_status = "Modified", "Matched", "Not Matched"
    prefix_characters = ['Mr', 'mr', 'Mr.', 'MR', 'mr.', 'MR.']
    dicitinory = {1: {'5': 'S'}, 2: {'S': '5'}, 3: {'0': 'O'}, 4: {'O': '0'}, 5: {'2': 'Z'}, 6: {'Z': '2'},
                  7: {'B': '8'}, 8: {'8': 'B'}, 9: {'1': 'I'}, 10: {'I': '1'}, 11: {'1': 'i'}, 12: {'i': '1'},
                  13: {'l': '1'}, 14: {'1': 'l'}, 15: {'1': 'L'}, 16: {'8': 'R'}, 17: {'R': '8'}, 18: {'Z': '7'},
                  19: {'7': 'Z'}, 20: {'I': 'l'}, 21: {'l': 'I'}, 22: {'O': 'Q'}, 23: {'Q': 'O'}, 24: {'L': 'I'},
                  25: {'I': 'L'}, 26: {'Q': 'G'}, 27: {'Q': 'g'}, 28: {'e': 'a'}, 30: {'a': 'e'}, 31: {'D': 'O'},
                  32: {'V': 'L'}}

    def dict_check(self):
        indeces_of_space = []
        idx = list(range(len(self.cheque_title)))
        for i in idx:
            if self.cheque_title[i] == ' ':
                indeces_of_space.append(i)
        cheque_title_list = list(re.sub(' +', '', self.cheque_title))
        acnt_title_list = list(re.sub(' +', '', self.acnt_title))
        if (len(acnt_title_list) == len(cheque_title_list)):
            for i in range(len(cheque_title_list)):
                if cheque_title_list[i].casefold() != acnt_title_list[i].casefold():
                    for key, value in self.dicitinory.items():
                        if acnt_title_list[i] == value.get(cheque_title_list[i]):
                            cheque_title_list[i] = value.get(cheque_title_list[i])
                            self.dict_flag = True
                            self.reason = "Found " + ''.join(
                                [k for k, v in value.items() if v == cheque_title_list[i]]) + " instead of " + \
                                          cheque_title_list[i] + " in cheque title"
        if cheque_title_list != acnt_title_list:
            self.dict_flag = False
            self.reason = " "
        for i in indeces_of_space:
            cheque_title_list.insert(i, ' ')
        self.cheque_title = ''.join(cheque_title_list)

    def regex_check(self, acnt_title_lower, cheque_title_lower):
        idx_dict = {}
        indeces_of_space = []
        spc_char_list = [' ','-', '.', ',', '/', '(', ')']
        special_characters = "[.,-/()]"
        acnt_title_list = list(self.acnt_title)
        for s in spc_char_list:
            for idx, val in enumerate(acnt_title_list):
                if s == val:
                    idx_dict.update({idx: val})
        idx_dict_list = list(idx_dict.keys())
        idx_dict_list.sort()
        idx_dict_sorted = {i: idx_dict[i] for i in idx_dict_list}
        idx = list(range(len(self.cheque_title)))
        acnt_title_without_sc = re.sub(special_characters, "", acnt_title_lower).replace(" ", "")
        cheque_title_without_sc = re.sub(special_characters, "", cheque_title_lower).replace(" ", "")
        acnt_title_without_sc_list = list(acnt_title_without_sc)
        cheque_title_without_sc_list = list(cheque_title_without_sc)
        acnts_title_list_without_space = list(re.sub(' +', '', self.acnt_title))
        cheque_title_list_without_space = list(re.sub(' +', '', self.cheque_title))
        if acnt_title_lower.__eq__(cheque_title_without_sc) or cheque_title_lower.__eq__(acnt_title_without_sc):

            self.status, self.ismatched = self.modified_status, True
            #for i in indeces_of_space:
             #   cheque_title_list_without_space.insert(i, ' ')
            if len(acnt_title_list) > len(cheque_title_list_without_space):
                print(idx_dict_sorted)
                for a in idx_dict_sorted.keys():
                    cheque_title_list_without_space.insert(a, idx_dict_sorted.get(a))
                if len(idx_dict_sorted) != 0:
                    self.reason = ''.join([a for a in list(idx_dict_sorted.values())]) + " is missing in cheque title"
            else:
                self.reason = ""
            final_chect_title = ''.join(cheque_title_list_without_space)
            idx_dict = {}
            return final_chect_title
        else:
            self.reason = " "
            self.dict_check()
            if (self.dict_flag == True):
                self.status, self.ismatched = self.modified_status, True
                self.dict_flag = False
                return self.cheque_title
            else:
                self.status, self.ismatched = self.Unsuccess_status, False
                return self.cheque_title

    def verify_title_excel(self):
        excel_file = openpyxl.load_workbook('Account_Cheque_Title_Matching.xlsx')
        # excel_file = openpyxl.load_workbook('test.xlsx')
        sheet = excel_file['Sheet1']
        rows = sheet.max_row
        print(rows)
        Acnt_title_col_idx, checque_title_col_idx, status_col_idx, reason_idx, Modified_chq_title_idx = 1, 2, 3, 4, 5
        sheet.cell(1, status_col_idx, "Status")
        sheet.cell(1, reason_idx, "Reason")
        sheet.cell(1, Modified_chq_title_idx, "Modified Cheque Title")
        for i in range(2, rows + 1):
            self.acnt_title = str(sheet.cell(i, Acnt_title_col_idx).value)
            self.cheque_title = str(sheet.cell(i, checque_title_col_idx).value)
            acnt_title_list = self.acnt_title.lower().split()
            cheque_title_list = self.cheque_title.lower().split()
            if self.acnt_title.casefold().__eq__(self.cheque_title.casefold()):
                self.reason = ""
                sheet.cell(i, status_col_idx, self.success_status)
                sheet.cell(i, reason_idx, self.reason)
                sheet.cell(i, Modified_chq_title_idx, self.cheque_title)
            else:
                cheque_title_after_regex = self.regex_check(self.acnt_title.casefold().replace(" ", ""),
                                                            self.cheque_title.casefold().replace(" ", ""))
                if self.ismatched:
                    sheet.cell(i, status_col_idx, self.status)
                    sheet.cell(i, reason_idx, self.reason)
                    sheet.cell(i, Modified_chq_title_idx, cheque_title_after_regex)
                else:
                    sheet.cell(i, status_col_idx, self.Unsuccess_status)
                    sheet.cell(i, reason_idx, self.reason)
                    sheet.cell(i, Modified_chq_title_idx, cheque_title_after_regex)
        # excel_file.save('test.xlsx')
        excel_file.save('Account_Cheque_Title_Matching.xlsx')
        excel_file.close()


if __name__ == '__main__':
    title_verification_obj = title_verification()
    title_verification_obj.verify_title_excel()
